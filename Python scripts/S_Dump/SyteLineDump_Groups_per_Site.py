#! Python 3.6.4
# Author: Alejandro Bautista R.

from openpyxl import load_workbook
from datetime import datetime
import openpyxl_styles


def write_to_workbook(*args):
    print("Opening existing workbook")
    # do not take into account the headers, so start in row 2
    row_cursor = 2

    # Load the existing workbook
    workbook = load_workbook(
        "C:\\Temp\\PythonOutputFiles\\Dump_" + str(datetime.now()).split(' ')[0] + ".xlsx")

    # Create a new worksheet
    worksheet = workbook.create_sheet()

    worksheet.title = "Groups_per_Site"

    # this line affects the first row
    worksheet.row_dimensions[1].height = 20

    # Titles for each column
    worksheet['A1'] = "Site"
    worksheet.column_dimensions['A'].width = 40
    worksheet.cell(row=1, column=1).font = openpyxl_styles.openpyxl_styles()

    worksheet['B1'] = "User"
    worksheet.column_dimensions['B'].width = 40
    worksheet.cell(row=1, column=2).font = openpyxl_styles.openpyxl_styles()

    worksheet['C1'] = "Group"
    worksheet.column_dimensions['C'].width = 40
    worksheet.cell(row=1, column=3).font = openpyxl_styles.openpyxl_styles()

    worksheet['D1'] = "Created"
    worksheet.column_dimensions['D'].width = 40
    worksheet.cell(row=1, column=4).font = openpyxl_styles.openpyxl_styles()

    # Insert values in each column #
    for index, value in enumerate(args[0]):
        try:
            worksheet.cell(row=row_cursor, column=1).value = value[0]
            worksheet.cell(row=row_cursor, column=2).value = value[1]
            worksheet.cell(row=row_cursor, column=3).value = value[2]
            worksheet.cell(row=row_cursor, column=4).value = value[3]
        except:
            print("Error in line %s\n data=%s" % (index, value))

        # Counter for going to the next row
        row_cursor = row_cursor + 1

    # save the workbook
    workbook.save(
        "C:\\Temp\\PythonOutputFiles\\Dump_" + str(datetime.now()).split(' ')[0] + ".xlsx")
    print("The results in the workbook have been saved.")

