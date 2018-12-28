#! Python 3.6.4
# Author: Alejandro Bautista R.

from openpyxl.workbook import Workbook
from datetime import datetime
import openpyxl_styles


def write_to_workbook(*args):
    print("Writing results to the Excel file")
    # do not take into account the headers, so start in row 2
    row_cursor = 2

    # Workbook headers
    workbook = Workbook()

    worksheet = workbook.create_sheet()

    worksheet.title = "Groups"

    # this line affects the first row
    worksheet.row_dimensions[1].height = 20

    # Titles for each column
    worksheet['A1'] = "GroupName"
    worksheet.column_dimensions['A'].width = 40
    worksheet.cell(row=1, column=1).font = openpyxl_styles.openpyxl_styles()

    worksheet['B1'] = "Group Desc"
    worksheet.column_dimensions['B'].width = 40
    worksheet.cell(row=1, column=2).font = openpyxl_styles.openpyxl_styles()

    worksheet['C1'] = "Form Name"
    worksheet.column_dimensions['C'].width = 40
    worksheet.cell(row=1, column=3).font = openpyxl_styles.openpyxl_styles()

    worksheet['D1'] = "Delete Privilege"
    worksheet.column_dimensions['D'].width = 40
    worksheet.cell(row=1, column=4).font = openpyxl_styles.openpyxl_styles()

    worksheet['E1'] = "Edit Privilege"
    worksheet.column_dimensions['E'].width = 40
    worksheet.cell(row=1, column=5).font = openpyxl_styles.openpyxl_styles()

    worksheet['F1'] = "Execute Privilege"
    worksheet.column_dimensions['F'].width = 40
    worksheet.cell(row=1, column=6).font = openpyxl_styles.openpyxl_styles()

    worksheet['G1'] = "Insert Privilege"
    worksheet.column_dimensions['G'].width = 40
    worksheet.cell(row=1, column=7).font = openpyxl_styles.openpyxl_styles()

    worksheet['H1'] = "Read Privilege"
    worksheet.column_dimensions['H'].width = 40
    worksheet.cell(row=1, column=8).font = openpyxl_styles.openpyxl_styles()

    worksheet['I1'] = "Update Privilege"
    worksheet.column_dimensions['I'].width = 40
    worksheet.cell(row=1, column=9).font = openpyxl_styles.openpyxl_styles()

    worksheet['J1'] = "Bulk Update Privilege"
    worksheet.column_dimensions['J'].width = 40
    worksheet.cell(row=1, column=10).font = openpyxl_styles.openpyxl_styles()

    # Insert values in each column #
    for index, value in enumerate(args[0]):
        try:
            worksheet.cell(row=row_cursor, column=1).value = value[0]
            worksheet.cell(row=row_cursor, column=2).value = value[1]
            worksheet.cell(row=row_cursor, column=3).value = value[2]
            worksheet.cell(row=row_cursor, column=4).value = value[3]
            worksheet.cell(row=row_cursor, column=5).value = value[4]
            worksheet.cell(row=row_cursor, column=6).value = value[5]
            worksheet.cell(row=row_cursor, column=7).value = value[6]
            worksheet.cell(row=row_cursor, column=8).value = value[7]
            worksheet.cell(row=row_cursor, column=9).value = value[8]
            worksheet.cell(row=row_cursor, column=10).value = value[9]
        except:
            print("Error in line %s\n data=%s" % (index, value))

        # Counter for going to the next row
        row_cursor = row_cursor + 1

    # delete the sheet one in Openpyxl because you do not want to display this empty sheet
    sheet_one = workbook['Sheet']
    workbook.remove(sheet_one)

    # save the workbook
    workbook.save(
        "C:\\Temp\\PythonOutputFiles\\Dump_" + str(datetime.now()).split(' ')[0] + ".xlsx")
    print("The results in the workbook have been saved.")

